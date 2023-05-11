import datetime
import os.path
import time
import traceback

import openpyxl
from excel_util import copy_cell_all
from openpyxl.utils import get_column_letter, column_index_from_string
from random_util import RandomUtil

def copy_row(worksheet, column_from, column_to, row_from, row_to):
    for column in range(column_index_from_string(column_from), column_index_from_string(column_to) + 1):
        cell_target = f"{get_column_letter(column)}{row_to}"
        cell_source = f"{get_column_letter(column)}{row_from}"
        copy_cell_all(worksheet[cell_source], worksheet[cell_target])

def generate_poster(poster_template_file, output_path, child_skus, parent_folder_name, error, finish):
    try:
        return_data = {}
        start_time = time.time()

        workbook = openpyxl.load_workbook(filename=poster_template_file)
        worksheet = workbook.active

        row_to_copy = 4
        max_column = get_column_letter(worksheet.max_column);
        print(max_column)
        for row, child_sku in enumerate(child_skus, start=4):
            worksheet[f"A{row}"] = child_sku
            copy_row(worksheet, "B", max_column, row_to_copy, row)

        workbook.close()
        end_time = time.time();

        now = datetime.datetime.now()
        current_datetime = now.strftime("%d%m%Y")
        saved_path = f"{output_path}/poster-{os.path.basename(parent_folder_name)}-{current_datetime}-{RandomUtil.random_digits(3)}.xlsx"
        workbook.save(saved_path)

        finish(end_time - start_time, saved_path, return_data)
    except Exception as e:
        traceback.print_exc()
        error(e)
