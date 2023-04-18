import argparse
import datetime
import os
import random
import uuid

import coloredlogs
import openpyxl
from random_util import RandomUtil
from s3_service import S3Service

parser = argparse.ArgumentParser("Upload tool")

parser.add_argument("-f", "--folder", type=str, help="Folder product")
parser.add_argument("-c", "--child", type=str, default="child", help="Folder child product images")
parser.add_argument("-t", "--title", type=str, default="Title.txt", help="Title file name")
parser.add_argument("-k", "--keyword", type=str, default="Keyword.txt", help="Keyword file name")
parser.add_argument("-p", "--price", type=str, default="Price.txt", help="Price file name")
parser.add_argument("-w", "--wall", type=str, default="wall.xlsx", help="Price file name")
parser.add_argument("-o", "--output", type=str, help="Output folder path")
args = parser.parse_args();

coloredlogs.install(level='INFO')

# Init S3
bucket_name = "ntnt-upload"
s3_service = S3Service(bucket_name)

# 1. Read images from given folder and "child folder" then rename image to UUID
folder_path = args.folder # "D:/Personal/MMO/Job/upload-tool/mountain 4"
child_folder_name = args.child
title_file = args.title
keyword_file = args.keyword
price_file = args.price
wall_file = args.wall
output_path = args.output

def rename_images_to_uuid(folder_path: str):
    for filename in os.listdir(folder_path):
        if filename.endswith((".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp")):
            image_path = os.path.join(folder_path, filename)
            new_filename = str(uuid.uuid4()) + os.path.splitext(filename)[1]
            new_image_path = os.path.join(folder_path, new_filename)
            os.rename(image_path, new_image_path)


rename_images_to_uuid(folder_path)
rename_images_to_uuid("{0}/{1}".format(folder_path, child_folder_name))

# 2. Get all main images from first level of that folder then upload to s3, then fill the image url into execel result file
## 2.1 Upload main images to S3
workbook = openpyxl.load_workbook(filename="result_excel.xlsx")
worksheet = workbook.active

main_image_urls = [None]
for filename in os.listdir(folder_path):
    if filename.endswith((".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp")):
        image_path = os.path.join(folder_path, filename)

        # Upload main image to S3
        image_s3_key = os.path.basename(folder_path) + "/" + filename
        s3_service.upload(image_path, image_s3_key)

        image_url = f"https://{bucket_name}.s3.amazonaws.com/{image_s3_key}"
        main_image_urls.append(image_url);

## 2.2 Fill main image data to excel file
start_row = 4
for i, value in enumerate(main_image_urls, start=start_row):
    cell = f"Q{i}"
    worksheet[cell] = value

# 3. Get all images from "child" folder then upload to s3 then fill the image url into result excel file
child_folder_path = folder_path + "/" + child_folder_name;
child_image_urls = []
for filename in os.listdir(child_folder_path):
    if filename.endswith((".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp")):
        image_path = os.path.join(child_folder_path, filename)

        # Upload main image to S3
        image_s3_key = os.path.basename(folder_path) + "/" + child_folder_name + "/" + filename
        s3_service.upload(image_path, image_s3_key)

        image_url = f"https://{bucket_name}.s3.amazonaws.com/{image_s3_key}"
        child_image_urls.append(image_url);

for position, letter in enumerate(["R", "S", "T", "U"], start=0):
    for i in range(start_row, start_row + len(main_image_urls)):
        cell = f"{letter}{i}"
        if (position < len(child_image_urls)):
            worksheet[cell] = child_image_urls[position]


# 4. Get information from "Title", "Keyword" and "Price". Then fill it into result excel file.

def add_multiple_rows(ws, column_letter: str, number_of_rows: int, start_row: int, get_each_row_data_func,
                      skipped_first_row=False):
    for i in range(0, number_of_rows):
        if (i == 0 and skipped_first_row):
            continue
        cell = f"{column_letter}{i + start_row}"
        ws[cell] = get_each_row_data_func(i)


with open(folder_path + "/" + title_file, "r") as f:
    title = f.read();
    add_multiple_rows(worksheet, "D", len(main_image_urls), start_row, lambda i: title + RandomUtil.random_digits(6))

with open(folder_path + "/" + keyword_file, "r") as f:
    keyword = f.read();
    add_multiple_rows(worksheet, "AB", len(main_image_urls), start_row, lambda i: keyword, True)

with open(folder_path + "/" + price_file, "r") as f:
    price = f.read();
    add_multiple_rows(worksheet, "O", len(main_image_urls), start_row, lambda i: float(price), True)

# 5. Read all information from "wall" files to calculate then fill it into result excel file.
workbook_wall = openpyxl.load_workbook(filename=folder_path + "/" + wall_file)
worksheet_wall = workbook_wall.active

ignored_first_row_columns = ["J", "K", "M", "P", "AI", "AJ"]
data_row_letters = ["A", "B", "C", "F", "G", "H", "I", "J", "K", "L", "M", "N", "P", "V", "W", "X", "Y", "Z", "AA",
                    "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR",
                    "AS"]
data_row_no = 2
for letter in data_row_letters:
    value = worksheet_wall[f"{letter}{data_row_no}"].value
    if (letter == "B"):
        prefix = value + "-"
        add_multiple_rows(worksheet, letter, len(main_image_urls), start_row,
                          lambda i: prefix + RandomUtil.random_string(10))
    elif (letter == "J"):
        prefix = value + " "
        add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, lambda i: prefix + "{:02d}".format(i),
                          True)
    elif (letter == "P"):
        add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, lambda i: random.randint(100, 999), True)
    elif (letter == "AH"):
        add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, lambda i: "Parent" if i == 0 else "Child")
    elif (letter == "AI"):
        sku = worksheet[f"B{start_row}"].value
        add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, lambda i: sku, True)
    else:
        add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, lambda i: value,
                          letter in ignored_first_row_columns)

# 7. End
now = datetime.datetime.now()
current_datetime = now.strftime("%d%m%Y-%H%M%S-%f")

workbook.save(f"{output_path}/exported_excel_{os.path.basename(folder_path)}_{current_datetime}-{RandomUtil.random_digits(3)}.xlsx")

workbook.close()
workbook_wall.close()
