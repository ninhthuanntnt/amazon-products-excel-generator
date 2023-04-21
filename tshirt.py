import datetime
import os
import random

import openpyxl
from random_util import RandomUtil
from s3_service import S3Service

def generate_t_shirt_excel(folder_path, output_path, error, finish):
    try:
        return_data = {"child_skus": []}
        ########## Init S3
        bucket_name = "ntnt-upload"
        s3_service = S3Service(bucket_name)

        ########## 1. Read images from given folder and "child folder" then rename image to UUID
        child_folder_name = "child"
        title_file = "Title.txt"
        keyword_file = "Keyword.txt"
        price_file = "Price.txt"
        data_file = "tshirt.xlsx"

        # def rename_images_to_uuid(folder_path: str):
        #     for filename in os.listdir(folder_path):
        #         if filename.endswith((".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp")):
        #             image_path = os.path.join(folder_path, filename)
        #             new_filename = str(uuid.uuid4()) + os.path.splitext(filename)[1]
        #             new_image_path = os.path.join(folder_path, new_filename)
        #             os.rename(image_path, new_image_path)
        #
        #
        # rename_images_to_uuid(folder_path)
        # rename_images_to_uuid("{0}/{1}".format(folder_path, child_folder_name))

        ########## 2. Get all main images from first level of that folder then upload to s3, then fill the image url into execel result file
        #################### 2.1 Upload main images to S3
        workbook = openpyxl.load_workbook(filename="tshirt_result_excel.xlsx")
        worksheet = workbook.active

        main_image_urls = [None]
        for filename in os.listdir(folder_path):
            if filename.endswith((".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp")):
                image_path = os.path.join(folder_path, filename)

                # Upload main image to S3
                image_s3_key = os.path.basename(folder_path) + "/" + filename
                s3_service.upload(image_path, image_s3_key)

                image_url = f"https://{bucket_name}.s3.amazonaws.com/{image_s3_key}"
                main_image_urls.append(image_url);

        #################### 2.2 Fill main image data to excel file
        start_row = 4
        for i, value in enumerate(main_image_urls, start=start_row):
            cell = f"Q{i}"
            worksheet[cell] = value

        ########## 3. Get all images from "child" folder then upload to s3 then fill the image url into result excel file
        child_folder_path = folder_path + "/" + child_folder_name;
        child_image_urls = []
        for filename in os.listdir(child_folder_path):
            if filename.endswith((".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp")):
                image_path = os.path.join(child_folder_path, filename)

                ########## Upload main image to S3
                image_s3_key = os.path.basename(folder_path) + "/" + child_folder_name + "/" + filename
                s3_service.upload(image_path, image_s3_key)

                image_url = f"https://{bucket_name}.s3.amazonaws.com/{image_s3_key}"
                child_image_urls.append(image_url);

        for position, letter in enumerate(["R", "S", "T", "U"], start=0):
            for i in range(start_row, start_row + len(main_image_urls)):
                cell = f"{letter}{i}"
                if (position < len(child_image_urls)):
                    worksheet[cell] = child_image_urls[position]

        ########## 4. Get information from "Title", "Keyword" and "Price". Then fill it into result excel file.

        def add_multiple_rows(ws, column_letter: str, number_of_rows: int, start_row: int, get_each_row_data_func,
                              skipped_first_row=False):
            for i in range(0, number_of_rows):
                if (i == 0 and skipped_first_row):
                    continue
                cell = f"{column_letter}{i + start_row}"
                ws[cell] = get_each_row_data_func(i)

        with open(folder_path + "/" + title_file, "r") as f:
            title = f.read();
            add_multiple_rows(worksheet, "D", len(main_image_urls), start_row,
                              lambda i: title + " " + RandomUtil.random_digits(6))

        with open(folder_path + "/" + keyword_file, "r") as f:
            keyword = f.read();
            add_multiple_rows(worksheet, "AB", len(main_image_urls), start_row, lambda i: keyword, True)

        with open(folder_path + "/" + price_file, "r") as f:
            price = f.read();
            add_multiple_rows(worksheet, "O", len(main_image_urls), start_row, lambda i: float(price), True)

        ########## 5. Read all information from "wall" files to calculate then fill it into result excel file.
        workbook_wall = openpyxl.load_workbook(filename=folder_path + "/" + data_file)
        worksheet_wall = workbook_wall.active

        ignored_first_row_columns = ["J", "K", "M", "P", "AI", "AJ"]
        data_row_letters = ["A", "B", "C", "F", "G", "H", "I", "J", "K", "L", "M", "N", "P", "V", "W", "X", "Y", "Z",
                            "AA", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP",
                            "AQ", "AR", "AS"]
        data_row_no = 2
        for letter in data_row_letters:
            value = worksheet_wall[f"{letter}{data_row_no}"].value
            ########## Handle SKU
            if (letter == "B"):
                prefix = value + "-" + os.path.basename(folder_path) + "-"

                def get_sku(i):
                    sku = prefix
                    if i == 0:
                        sku = prefix + srg.generator.generate_from_milisecond() + "-" + RandomUtil.random_string(3)
                    else:
                        sku = prefix + os.path.splitext(main_image_filenames[i])[0] + "-" + RandomUtil.random_string(3)
                        return_data["child_skus"].append(sku)
                    return sku

                add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, get_sku)
                return_data["child_skus"].append(sku)
            ########## Handle color name with value + autoincrement number
            elif (letter == "J"):
                prefix = value + " "
                add_multiple_rows(worksheet, letter, len(main_image_urls), start_row,
                                  lambda i: prefix + "{:02d}".format(i),
                                  True)
            ########## Quantity
            elif (letter == "P"):
                add_multiple_rows(worksheet, letter, len(main_image_urls), start_row,
                                  lambda i: random.randint(100, 999), True)
            ########## Parent child
            elif (letter == "AH"):
                add_multiple_rows(worksheet, letter, len(main_image_urls), start_row,
                                  lambda i: "Parent" if i == 0 else "Child")
            ########## Parent SKU
            elif (letter == "AI"):
                sku = worksheet[f"B{start_row}"].value
                add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, lambda i: sku, True)
            else:
                add_multiple_rows(worksheet, letter, len(main_image_urls), start_row, lambda i: value,
                                  letter in ignored_first_row_columns)

        ########## 7. End
        now = datetime.datetime.now()
        current_datetime = now.strftime("%d%m%Y")
        path_to_save = f"{output_path}/{os.path.basename(folder_path)}_{current_datetime}-{RandomUtil.random_digits(3)}.xlsx"
        workbook.save(path_to_save)

        workbook.close()
        workbook_wall.close()

        end_time = time.time();
        finish(end_time - start_time, path_to_save, return_data)
    except Exception as e:
        print(f"Error when upload folder: {folder_path}")
        traceback.print_exc()
        error(e)
